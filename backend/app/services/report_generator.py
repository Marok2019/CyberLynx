from datetime import datetime
from io import BytesIO, StringIO
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv

class ReportGenerator:
    """Generador de reportes en múltiples formatos para auditorías"""
    
    @staticmethod
    def generate_pdf_report(audit, checklist_data):
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TituloPortada',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=HexColor('#1976d2'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'TituloSeccion',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=HexColor('#333333'),
            spaceAfter=12
        )

        story = []

        # === Portada (página 1) ===
        story.append(Spacer(1, 2*inch))
        story.append(Paragraph("CyberLynx", title_style))
        story.append(Paragraph("Reporte de Auditoría de Seguridad", styles['Heading2']))
        story.append(Spacer(1, 0.5*inch))

        audit_info = f"""
        <b>Nombre de la auditoría:</b> {audit.name}<br/>
        <b>Estado:</b> {audit.status}<br/>
        <b>Fecha de inicio:</b> {audit.created_at.strftime('%d/%m/%Y %H:%M')}<br/>
        <b>Fecha de cierre:</b> {audit.completed_at.strftime('%d/%m/%Y %H:%M') if audit.completed_at else 'En progreso'}<br/>
        """
        story.append(Paragraph(audit_info, styles['Normal']))

        if audit.description:
            story.append(Spacer(1, 0.3*inch))
            story.append(Paragraph(f"<b>Descripción:</b><br/>{audit.description}", styles['Normal']))

        story.append(PageBreak())

        # === Resumen Ejecutivo (página 2) ===
        story.append(Paragraph("Resumen Ejecutivo", heading_style))

        total_questions = 0
        total_yes = 0
        total_no = 0
        total_na = 0
        critical_findings = []
        high_findings = []

        for checklist in checklist_data:
            summary = checklist['summary']
            total_questions += summary['total_questions']
            total_yes += summary['yes_count']
            total_no += summary['no_count']
            total_na += summary['na_count']

            for severity, stats in summary['severity_breakdown'].items():
                if severity == 'Critical' and stats['no'] > 0:
                    critical_findings.append({
                        'checklist': checklist['name'],
                        'count': stats['no']
                    })
                elif severity == 'High' and stats['no'] > 0:
                    high_findings.append({
                        'checklist': checklist['name'],
                        'count': stats['no']
                    })

        compliance_rate = round((total_yes / (total_yes + total_no) * 100), 2) if (total_yes + total_no) > 0 else 0

        resumen_texto = f"""
        <b>Preguntas evaluadas:</b> {total_questions}<br/>
        <b>Cumple (Sí):</b> {total_yes}<br/>
        <b>No cumple (No):</b> {total_no}<br/>
        <b>No aplica (N/A):</b> {total_na}<br/>
        <b>Porcentaje de cumplimiento global:</b> {compliance_rate}%<br/><br/>
        <b>Hallazgos críticos:</b> {len(critical_findings)}<br/>
        <b>Hallazgos de prioridad alta:</b> {len(high_findings)}<br/>
        """
        story.append(Paragraph(resumen_texto, styles['Normal']))

        story.append(PageBreak())

        # === Detalle por Checklist (cada bloque se mantiene junto) ===
        story.append(Paragraph("Detalle por Checklist", heading_style))

        sev_map = {'Critical': 'Crítica', 'High': 'Alta', 'Medium': 'Media', 'Low': 'Baja'}

        for checklist in checklist_data:
            bloque = []
            bloque.append(Spacer(1, 0.2*inch))
            bloque.append(
                Paragraph(
                    f"<b>{checklist['name']}</b> ({checklist['category']})",
                    styles['Heading3']
                )
            )

            summary = checklist['summary']
            severity_data = [['Severidad', 'Total', 'Sí', 'No', 'N/A', 'Sin responder']]

            for sev in ['Critical', 'High', 'Medium', 'Low']:
                if sev in summary['severity_breakdown']:
                    stats = summary['severity_breakdown'][sev]
                    severity_data.append([
                        sev_map.get(sev, sev),
                        stats['total'],
                        stats['yes'],
                        stats['no'],
                        stats['na'],
                        stats['unanswered']
                    ])

            severity_table = Table(severity_data, colWidths=[1.3*inch]*6)
            severity_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#1976d2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), HexColor('#f5f5f5')),
                ('GRID', (0, 0), (-1, -1), 1, HexColor('#cccccc'))
            ]))

            bloque.append(severity_table)
            bloque.append(Spacer(1, 0.3*inch))

            story.append(KeepTogether(bloque))

        story.append(PageBreak())

        # === Hallazgos críticos ===
        if critical_findings or high_findings:
            story.append(Paragraph("Hallazgos Críticos y de Alta Prioridad", heading_style))
            hallazgos_data = [['Checklist', 'Severidad', 'Cantidad', 'Acción requerida']]

            for finding in critical_findings:
                hallazgos_data.append([
                    finding['checklist'], 'Crítica', finding['count'],
                    'Atender de inmediato'
                ])
            for finding in high_findings:
                hallazgos_data.append([
                    finding['checklist'], 'Alta', finding['count'],
                    'Resolver en menos de 30 días'
                ])

            hallazgos_table = Table(hallazgos_data, colWidths=[2.5*inch, 1*inch, 0.8*inch, 2*inch])
            hallazgos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#d32f2f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, HexColor('#cccccc')),
                ('BACKGROUND', (0, 1), (-1, -1), HexColor('#ffebee'))
            ]))
            story.append(hallazgos_table)

            story.append(PageBreak())

        # === Recomendaciones ===
        story.append(Paragraph("Recomendaciones", heading_style))
        recomendaciones = [
            "1. Atienda inmediatamente los hallazgos críticos identificados.",
            "2. Genere un plan de acción para hallazgos de prioridad alta dentro de 30 días.",
            "3. Programe auditorías periódicas para dar seguimiento a la mejora continua.",
            "4. Refuerce la capacitación en seguridad para todo el personal.",
            "5. Actualice políticas de seguridad conforme a las brechas detectadas."
        ]
        for rec in recomendaciones:
            story.append(Paragraph(rec, styles['Normal']))
            story.append(Spacer(1, 0.1*inch))

        story.append(Spacer(1, 0.5*inch))
        footer_text = f"<i>Reporte generado por CyberLynx el {datetime.now().strftime('%d/%m/%Y %H:%M')}</i>"
        story.append(Paragraph(footer_text, styles['Normal']))

        doc.build(story)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_excel_report(audit, checklist_data):
        buffer = BytesIO()
        wb = Workbook()

        ws_summary = wb.active
        ws_summary.title = "Resumen"
        ws_summary['A1'] = "Reporte de Auditoría de Seguridad - CyberLynx"
        ws_summary['A1'].font = Font(size=16, bold=True, color="1976D2")
        ws_summary.merge_cells('A1:D1')

        ws_summary['A3'] = "Nombre de la auditoría:"
        ws_summary['B3'] = audit.name
        ws_summary['A4'] = "Estado:"
        ws_summary['B4'] = audit.status
        ws_summary['A5'] = "Fecha de inicio:"
        ws_summary['B5'] = audit.created_at.strftime('%d/%m/%Y %H:%M')
        ws_summary['A6'] = "Fecha de finalización:"
        ws_summary['B6'] = audit.completed_at.strftime('%d/%m/%Y %H:%M') if audit.completed_at else 'En progreso'
        